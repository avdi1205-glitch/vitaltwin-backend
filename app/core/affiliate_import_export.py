"""Affiliate product import/export — CSV, JSON, and Excel (.xlsx).

VitalTwin Enterprise Release — Affiliate Intelligence & Management Platform.

Real parsing/serialization only (stdlib `csv`/`json` + `openpyxl` for
Excel) — no vendor "connector" of any kind. Import resolves `category`/
`partner` *names* to their existing ids on a best-effort basis and reports
every row that could not be resolved or was otherwise invalid, instead of
silently dropping or guessing data.
"""

from __future__ import annotations

import csv
import io
import json

from openpyxl import Workbook, load_workbook

from .supabase import supabase

PRODUCT_TABLE = "vt_affiliate_products"
CATEGORY_TABLE = "vt_affiliate_categories"
PARTNER_TABLE = "vt_affiliate_partners"

EXPORT_FIELDS = [
    "id", "title", "subtitle", "category", "brand", "manufacturer", "description",
    "image_url", "price", "currency", "affiliate_url", "deep_link", "partner",
    "commission_rate", "tags", "target_audience", "region", "language", "status",
    "priority", "rating", "notes", "start_date", "end_date",
]

IMPORT_FIELDS = [f for f in EXPORT_FIELDS if f != "id"]


def _category_lookup() -> dict[str, str]:
    try:
        rows = supabase.table(CATEGORY_TABLE).select("id,name").execute().data or []
    except Exception:
        return {}
    return {row["name"]: row["id"] for row in rows}


def _category_names_by_id() -> dict[str, str]:
    return {v: k for k, v in _category_lookup().items()}


def _partner_lookup() -> dict[str, str]:
    try:
        rows = supabase.table(PARTNER_TABLE).select("id,partner_name").execute().data or []
    except Exception:
        return {}
    return {row["partner_name"]: row["id"] for row in rows if row.get("partner_name")}


def _partner_names_by_id() -> dict[str, str]:
    return {v: k for k, v in _partner_lookup().items()}


def export_products(fmt: str) -> tuple[bytes, str, str]:
    """Returns `(content_bytes, media_type, filename)`."""
    try:
        rows = supabase.table(PRODUCT_TABLE).select("*").order("created_at", desc=True).execute().data or []
    except Exception:
        rows = []

    category_names = _category_names_by_id()
    partner_names = _partner_names_by_id()
    records = []
    for row in rows:
        record = {field: row.get(field) for field in EXPORT_FIELDS if field != "category" and field != "partner"}
        record["id"] = row.get("id")
        record["category"] = category_names.get(str(row.get("category_id")), "")
        record["partner"] = partner_names.get(str(row.get("partner_id")), "")
        record["tags"] = ";".join(row.get("tags") or [])
        records.append(record)

    if fmt == "json":
        content = json.dumps(records, ensure_ascii=False, indent=2, default=str).encode("utf-8")
        return content, "application/json", "affiliate_products.json"

    if fmt == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Produkte"
        sheet.append(EXPORT_FIELDS)
        for record in records:
            sheet.append([str(record.get(field, "") or "") for field in EXPORT_FIELDS])
        buffer = io.BytesIO()
        workbook.save(buffer)
        return (
            buffer.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "affiliate_products.xlsx",
        )

    # default: csv
    buffer_str = io.StringIO()
    writer = csv.DictWriter(buffer_str, fieldnames=EXPORT_FIELDS)
    writer.writeheader()
    for record in records:
        writer.writerow(record)
    return buffer_str.getvalue().encode("utf-8"), "text/csv", "affiliate_products.csv"


def _rows_from_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))


def _rows_from_json(content: bytes) -> list[dict]:
    parsed = json.loads(content.decode("utf-8"))
    if isinstance(parsed, dict):
        parsed = parsed.get("items", [])
    if not isinstance(parsed, list):
        raise ValueError("JSON muss eine Liste von Produkten sein.")
    return parsed


def _rows_from_xlsx(content: bytes) -> list[dict]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    header = [str(cell or "").strip() for cell in next(rows_iter, [])]
    rows = []
    for values in rows_iter:
        row = dict(zip(header, values))
        if any(v not in (None, "") for v in row.values()):
            rows.append({k: ("" if v is None else v) for k, v in row.items()})
    return rows


def parse_import_file(fmt: str, content: bytes) -> list[dict]:
    if fmt == "json":
        return _rows_from_json(content)
    if fmt == "xlsx":
        return _rows_from_xlsx(content)
    return _rows_from_csv(content)


def import_products(fmt: str, content: bytes, *, created_by: str) -> dict:
    rows = parse_import_file(fmt, content)
    category_ids = _category_lookup()
    partner_ids = _partner_lookup()

    imported = 0
    errors: list[dict] = []
    for index, row in enumerate(rows, start=1):
        title = str(row.get("title") or "").strip()
        affiliate_url = str(row.get("affiliate_url") or "").strip()
        if not title or not affiliate_url:
            errors.append({"row": index, "error": "title und affiliate_url sind Pflichtfelder."})
            continue

        category_name = str(row.get("category") or "").strip()
        partner_name = str(row.get("partner") or "").strip()
        category_id = category_ids.get(category_name) if category_name else None
        partner_id = partner_ids.get(partner_name) if partner_name else None
        if category_name and not category_id:
            errors.append({"row": index, "error": f"Kategorie '{category_name}' nicht gefunden — ohne Kategorie importiert."})
        if partner_name and not partner_id:
            errors.append({"row": index, "error": f"Partnerprogramm '{partner_name}' nicht gefunden — ohne Partner importiert."})

        tags_raw = row.get("tags") or ""
        tags = [t.strip() for t in str(tags_raw).split(";") if t.strip()]

        payload = {
            "title": title,
            "subtitle": row.get("subtitle") or None,
            "category_id": category_id,
            "brand": row.get("brand") or None,
            "manufacturer": row.get("manufacturer") or None,
            "description": row.get("description") or None,
            "image_url": row.get("image_url") or None,
            "price": row.get("price") or None,
            "currency": row.get("currency") or "eur",
            "affiliate_url": affiliate_url,
            "deep_link": row.get("deep_link") or None,
            "partner_id": partner_id,
            "commission_rate": row.get("commission_rate") or None,
            "tags": tags,
            "target_audience": row.get("target_audience") or None,
            "region": row.get("region") or "DE",
            "language": row.get("language") or "de",
            "status": row.get("status") or "draft",
            "priority": int(row.get("priority") or 0),
            "rating": row.get("rating") or None,
            "notes": row.get("notes") or "",
            "start_date": row.get("start_date") or None,
            "end_date": row.get("end_date") or None,
            "created_by": created_by,
        }
        try:
            supabase.table(PRODUCT_TABLE).insert(payload).execute()
            imported += 1
        except Exception as exc:
            errors.append({"row": index, "error": str(exc)})

    return {"imported": imported, "total_rows": len(rows), "errors": errors}
